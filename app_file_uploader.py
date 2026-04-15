"""
基于Streamlit完成WEB网页上传服务

pip install streamlit

Streamlit：当WEB页面元素发生变化，则代码重新执行一遍
"""
import csv
import io
import json
import os
import time
import streamlit as st
from knowledge_base import KnowledgeBaseService

st.set_page_config(page_title="知识库更新", layout="centered")

st.markdown(
    """
    <style>
      /* 炫彩风格：高饱和渐变、光晕、动态效果 */
      .block-container { padding-top: 2.2rem; position: relative; }
      .block-container::before {
        content: "";
        position: fixed;
        inset: 0;
        pointer-events: none;
        z-index: -1;
        background:
          radial-gradient(48rem 26rem at 6% 12%, rgba(236, 72, 153, 0.18), transparent 60%),
          radial-gradient(45rem 24rem at 92% 6%, rgba(59, 130, 246, 0.16), transparent 60%),
          radial-gradient(50rem 25rem at 50% 100%, rgba(16, 185, 129, 0.14), transparent 60%),
          linear-gradient(145deg, #f8fafc 0%, #eef2ff 35%, #f5f3ff 65%, #ecfeff 100%);
      }

      .kb-hero {
        border-radius: 22px;
        padding: 16px 16px;
        margin-bottom: 14px;
        border: 1px solid rgba(99, 102, 241, 0.30);
        background:
          radial-gradient(880px 260px at 4% -12%, rgba(244, 63, 94, 0.22), transparent 60%),
          radial-gradient(760px 260px at 96% -10%, rgba(59, 130, 246, 0.23), transparent 60%),
          linear-gradient(125deg, rgba(255,255,255,0.92) 0%, rgba(238,242,255,0.92) 35%, rgba(245,243,255,0.92) 70%, rgba(236,253,245,0.92) 100%);
        box-shadow: 0 16px 38px rgba(99, 102, 241, 0.16);
      }
      .kb-hero-title {
        font-size: 1.25rem;
        font-weight: 800;
        letter-spacing: 0.2px;
        background: linear-gradient(90deg, #ef4444 0%, #f59e0b 22%, #22c55e 44%, #3b82f6 66%, #a855f7 100%);
        -webkit-background-clip: text;
        background-clip: text;
        -webkit-text-fill-color: transparent;
      }
      .kb-hero-sub { color: rgba(15, 23, 42, 0.72); margin-top: 4px; line-height: 1.55; }

      .kb-card {
        border: 1px solid rgba(99, 102, 241, 0.28);
        border-radius: 20px;
        padding: 16px 16px;
        background:
          linear-gradient(135deg, rgba(255,255,255,0.88) 0%, rgba(248,250,252,0.75) 100%);
        box-shadow: 0 14px 34px rgba(76, 29, 149, 0.14);
        transition: transform 160ms ease, box-shadow 160ms ease, border-color 160ms ease;
        backdrop-filter: blur(8px);
      }
      .kb-card:hover {
        transform: translateY(-2px) scale(1.003);
        box-shadow: 0 22px 46px rgba(79, 70, 229, 0.2);
        border-color: rgba(168, 85, 247, 0.45);
      }
      .kb-muted { color: rgba(15, 23, 42, 0.7); font-size: 0.95rem; }
      .kb-kv { display:flex; gap:12px; flex-wrap:wrap; margin-top: 6px; }
      .kb-pill {
        display:inline-flex;
        align-items:center;
        gap:6px;
        padding: 6px 10px;
        border-radius: 999px;
        border: 1px solid rgba(129, 140, 248, 0.45);
        background: linear-gradient(135deg, rgba(255,255,255,0.96) 0%, rgba(238,242,255,0.98) 46%, rgba(245,243,255,0.98) 100%);
        font-size: 0.9rem;
        box-shadow: 0 6px 14px rgba(79, 70, 229, 0.12);
      }
      .kb-hr { margin: 10px 0 14px 0; border: none; border-top: 1px solid rgba(148, 163, 184, 0.35); }
      code { font-size: 0.9em; }

      /* 上传区域与上传按钮 */
      [data-testid="stFileUploader"] > div {
        border-radius: 18px !important;
        border: 1px solid rgba(129, 140, 248, 0.45) !important;
        background: linear-gradient(135deg, rgba(255,255,255,0.86) 0%, rgba(238,242,255,0.86) 100%) !important;
        box-shadow: 0 10px 24px rgba(79, 70, 229, 0.12) !important;
      }
      [data-testid="stFileUploader"] small {
        color: rgba(79, 70, 229, 0.85) !important;
      }
      [data-testid="stFileUploader"] section button {
        border-radius: 999px !important;
        border: 0 !important;
        color: #ffffff !important;
        font-weight: 700 !important;
        letter-spacing: 0.2px;
        padding: 0.45rem 1rem !important;
        background: linear-gradient(120deg, #ec4899 0%, #f59e0b 35%, #22c55e 65%, #3b82f6 100%) !important;
        background-size: 190% 190% !important;
        box-shadow: 0 10px 22px rgba(79, 70, 229, 0.28) !important;
        animation: kb-shift 6s ease infinite;
      }
      [data-testid="stFileUploader"] section button:hover {
        transform: translateY(-1px);
        filter: brightness(1.08);
        box-shadow: 0 16px 30px rgba(79, 70, 229, 0.32) !important;
      }

      /* 按钮更精致一点（尽量不影响 Streamlit 默认） */
      div.stButton > button[kind="primary"] {
        border-radius: 12px !important;
        background: linear-gradient(120deg, #f43f5e 0%, #f59e0b 25%, #22c55e 50%, #3b82f6 75%, #8b5cf6 100%) !important;
        background-size: 200% 200% !important;
        border: 0 !important;
        color: #ffffff !important;
        box-shadow: 0 12px 28px rgba(99, 102, 241, 0.30);
        animation: kb-shift 6s ease infinite;
      }
      div.stButton > button[kind="primary"]:hover {
        transform: translateY(-1px);
        filter: brightness(1.08);
        box-shadow: 0 18px 34px rgba(79, 70, 229, 0.34);
      }
      div.stDownloadButton > button {
        border-radius: 12px !important;
        border: 1px solid rgba(129, 140, 248, 0.55) !important;
        background: linear-gradient(135deg, rgba(236, 72, 153, 0.14) 0%, rgba(59, 130, 246, 0.14) 100%) !important;
      }
      div.stDownloadButton > button:hover {
        border-color: rgba(236, 72, 153, 0.55) !important;
      }
      @keyframes kb-shift {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
      }
    </style>
    """,
    unsafe_allow_html=True,
)

type_list = ["txt", "md", "csv", "json", "pdf", "docx", "xlsx"]


def else_to_txt(file) -> str:#将上传的文件内容转化为文本内容
    last = os.path.splitext(file.name)[1].lower()#取后缀并统一转化为小写
    bin = file.getvalue()#把上传的文件读成二进制数据

    if last in {".txt", ".md"}:#文本直接返回
        return bin.decode("utf-8")

    if last == ".json":
        j = json.loads(file_bytes.decode("utf-8"))
        return json.dumps(j, ensure_ascii=False, indent=2)

    if last == ".csv":
        content = bin.decode("utf-8")#解码成文本
        reader = csv.reader(io.StringIO(content))#按行读取表格
        rows = ["\t".join(row) for row in reader]#每一行用制表符隔开，方便阅读
        return "\n".join(rows)#返回纯文本格式的表格内容

    if last == ".pdf":
        from pypdf import PdfRead
        reader = PdfReader(io.BytesIO(bin))#一页一页提取文字
        pages = [page.extract_text() or "" for page in reader.pages]#把所有页拼接起来
        return "\n".join(pages).strip()#返回纯文本

    if last == ".docx":
        from docx import Document
        doc = Document(io.BytesIO(bin))#读取所有段落文字
        return "\n".join(p.text for p in doc.paragraphs if p.text).strip()#过滤空行

    if last == ".xlsx":
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(bin), data_only=True)
        all_rows = []
        for sheet in workbook.worksheets:#遍历所有工作表
            all_rows.append(f"### Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):#逐行读取单元格内容
                row_values = ["" if v is None else str(v) for v in row]
                all_rows.append("\t".join(row_values))#用制表符排版，变成纯文本表格
        return "\n".join(all_rows).strip()

    raise ValueError(f"暂不支持的文件类型：{suffix}")#都不匹配，抛出错误


st.markdown(
    """
    <div class="kb-hero">
      <div class="kb-hero-title">知识库更新</div>
      <div class="kb-hero-sub">
        上传多种格式文件。支持 <code>.txt / .md / .csv / .json / .pdf / .docx / .xlsx</code>。
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# 文件上传（支持多选）
files = st.file_uploader(
    "请选择一个或多个知识文件",
    type=type_list,
    accept_multiple_files=True
)

# 初始化知识库服务
if "service" not in st.session_state:
    st.session_state["service"] = KnowledgeBaseService()

# 处理多文件上传
if files:
    total = len(files)
    st.success(f"已选择 {total} 个文件，开始加载知识库...")
    for idx, file in enumerate(files, 1):# 逐个处理文件
        with st.container():
            st.markdown("---")
            st.markdown(f"####  正在处理第 {idx}/{total} 个文件")
            file_name = file.name
            file_size_kb = file.size / 1024
            st.info(f"**文件名**: {file_name}  |  **大小**: {file_size_kb:.2f} KB")# 展示文件信息
            try:
                text = else_to_txt(file)# 读取文件并提取文本
            except Exception as e:
                st.error(f" {file_name} 解析失败：{e}")
                continue
            if not text.strip():
                st.warning(f"{file_name} 未提取到有效文本，跳过。")
                continue
            with st.spinner(f"正在载入 {file_name}..."):# 加载动画
                time.sleep(0.8)
                result = st.session_state["service"].upload_by_str(text, file_name)

            st.success(f"**{file_name} 上传完成**")# 成功提示
            with st.expander("查看解析详情", expanded=False):
                st.write(result)

    st.markdown("---")
    st.balloons()
    st.success("所有文件加载完毕")



